import pandas as pd
import openpyxl as px
from datetime import datetime
#output date format [mm/dd/yy]

def read_excel_and_create_dict(x):
    # Read the Excel file
    df = pd.read_excel(x)
    dff = px.load_workbook(x)
    sheet=dff.active

    # Create a dictionary to store the information
    result_dict = {}
    #print(sheet.cell(row=3,column=7).value)

    # Iterate over rows and extract information
    for index, row in df.iterrows():
        name = row['NAME']  # Assuming 'NAME' is the column header for names
        dates = []

        # Iterate over columns starting from the 6th column (Week 1)

        ff=6
        for col in df.columns[6:]:
            if row[col] == '✓':
                dts=sheet.cell(row=3,column=ff).value
                dtss=dts.strftime("%x")
                dates.append(dtss)
            ff=ff+1
            

        # Add to the dictionary only if there are dates with '✓'
        if dates:
            result_dict[name] = dates

    return result_dict

# Replace 'xl.xlsx' with the actual path to your Excel file
excel_file_path = 'ha_tracker.xlsx'
result_dictionary = read_excel_and_create_dict(excel_file_path)

# Print the resulting dictionary
print(result_dictionary)