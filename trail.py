import pandas as pd
import openpyxl as px
from datetime import datetime
#output date format [mm/dd/yy]

def read_excel_and_create_dict(x):
    # Read the Excel file
    df = pd.read_excel(x)
    dff = px.load_workbook(x)
    sheet=dff.active

    # Create a dictionary to store the information
    result_dict = {}
    #print(sheet.cell(row=3,column=7).value)

    # Iterate over rows and extract information
    for index, row in df.iterrows():
        name = row['NAME']  # Assuming 'NAME' is the column header for names
        dates = []

        # Iterate over columns starting from the 6th column (Week 1)

        ff=6
        for col in df.columns[6:]:
            if row[col] == '✓':
                dts=sheet.cell(row=3,column=ff).value
                dtss=dts.strftime("%x")
                dates.append(dtss)
            ff=ff+1
            

        # Add to the dictionary only if there are dates with '✓'
        if dates:
            result_dict[name] = dates

    return result_dict

# Replace 'xl.xlsx' with the actual path to your Excel file
excel_file_path = 'ha_tracker.xlsx'
result_dictionary = read_excel_and_create_dict(excel_file_path)

def ha(pt_taken): #mm dd yy
    from datetime import datetime
    def inputdate(nominal):
        a=pt_taken[nominal]
        j,k,l=a.split("/")
        dateIs = datetime(int(l),int(j),int(k))
        toOrdinal = dateIs.toordinal()
        return toOrdinal

    #exact same as the above function but used for date of commencement
    def inputcomm():
        a=pt_taken[0]
        j,k,l=a.split("/")
        dateIs = datetime(int(l),int(j),int(k))
        toOrdinal = dateIs.toordinal()
        return toOrdinal

    ha = 1 #assuming HA=true at the start 
    nominal=1
    x= inputcomm() #the first date example: enlistment date (date is inputed at numner as of now)
    y= inputdate(pt_taken[1])#first PT
    ha1=x+14
    while True:
        nominal=nominal+1
        if y>ha1:
            print("ha lost")
            break
        elif y<ha1 and ha==1:
            z=inputdate(pt_taken[nominal]) #second PT
            d=z-y #duration
            if z>ha1:
                print("ha lost")
                break
            elif d<=7 and z<ha1:
                print("ha true")
                ha1=z+14
                y=z
                continue
            elif d>7:
                ha = 2 #means pending
                print("ha pending")
                continue
            print("ha lost")
            break
        elif ha==2:
            a=inputdate(pt_taken[nominal]) #third PT
            if (a-z)<=7 and a<=ha1 :
                ha = 1
                print("ha true")
                y=a
                ha1=a+14
            else:
                print("ha lost")
                break        

for names in result_dictionary:
    pt_date = result_dictionary[names]
    ha(pt_date)

print("hello")